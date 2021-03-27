from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.urls import reverse, reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.generic import DetailView, ListView
from django.views.generic.edit import CreateView, UpdateView
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Count

from django.core.files.storage import default_storage

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors, PatternFill

import json
import os
import copy
import re
import pickle

import nltk
import string
from nltk.corpus import stopwords
from nltk.corpus import swadesh
from nltk.tokenize import word_tokenize
from nltk.tokenize import sent_tokenize
import markovify

from .models import Author, Writing, Sample, GENRE_CHOICES
from .forms import AuthorForm, WritingForm
from mysite import settings
from . import tools
from .tools import DISTANCE_FUNCTIONS, PLOT_FUNCTIONS

def index(request):
    stop_words = stopwords.words('russian')
    messages.info(request, f'STOP WORDS: {stop_words}', extra_tags='alert-warning')
    swadesh_words = swadesh.words('ru')
    messages.info(request, f'SWADESH WORDS: {swadesh_words}', extra_tags='alert-warning')
    return redirect(reverse('pushkin:author_list'))

INVERSE = None
DIRECT = None
OBJECT_COUNT = 0
PARAM_COUNT = 0
CONTINUE = False

def vector(direct, inverse) :
    v = []
    for label in inverse:
        item = dict()
        item['name'] = label
        item['percent'] = inverse[label]['percent']
        data = []
        for param in inverse[label]['data']:
            data_item = dict()
            data_item['param'] = param
            data_item['d_value'] = direct[label]['data'][param]
            data_item['i_value'] = inverse[label]['data'][param]
            data.append(data_item)
        item['data'] = data
        v.append(item)
    return v

# _______________________________________________________ AUTHOR _______________

class AuthorUpdate(LoginRequiredMixin, UpdateView) :
    model = Author
    form_class = AuthorForm
    success_url = reverse_lazy('pushkin:author_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

    def form_valid(self, form):
        return super().form_valid(form)


def author_delete(request, id=None):
    if not id:
        messages.info(request, "Delete, No ID", extra_tags='alert-warning')
    author = Author.objects.filter(id=id).first()
    if not author:
        messages.info(request, f"Delete, No {id} author", extra_tags='alert-warning')
        return redirect(reverse('pushkin:author_list'))
    author.delete()
    return redirect(reverse('pushkin:author_list'))


class AuthorCreate(LoginRequiredMixin, CreateView) :
    model = Author
    form_class = AuthorForm
    success_url = reverse_lazy('pushkin:author_list')

    def form_valid(self, form):
        if not os.path.exists(form.instance.writings_dir()):
            os.mkdir(form.instance.writings_dir())
        return super().form_valid(form)


class AuthorList(LoginRequiredMixin, ListView) :
    model = Author

    def get_queryset(self, **kwargs):
        mode = None
        if 'mode' in self.kwargs:
            mode = self.kwargs['mode']
        if mode == 'checked_only':
            authors = Author.objects.filter(active=True)
        else:
            authors = Author.objects.all()
        return authors

    def get_context_data(self, **kwargs):
        global INVERSE
        global DIRECT
        global OBJECT_COUNT
        global PARAM_COUNT
        global CONTINUE
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)

        INVERSE = None
        DIRECT = None
        OBJECT_COUNT = 0
        PARAM_COUNT = 0
        CONTINUE = False

        mode = None
        if 'mode' in self.kwargs:
            mode = self.kwargs['mode']

        writings = Writing.objects.all()

        if mode == 'checked_only':
            writings = writings.filter(active=True)

        context['writings'] = writings
        context['active_a'] = Author.objects.filter(active=True).count()
        context['active_w'] = Writing.objects.filter(active=True).count()
        context['genres'] = GENRE_CHOICES
        return context

def get_detail_context(request, context, obj, objects):
    global INVERSE
    global DIRECT
    global OBJECT_COUNT
    global PARAM_COUNT
    global CONTINUE

    N = objects.count()
    if N == 0 :
        messages.info(request, f"У {obj} нет выбранных объектов для сравнения, выберите хотя-бы один.", extra_tags='alert-danger')
        return context

    if not CONTINUE: # First call
        DIRECT = dict()
        INVERSE = dict()
        for object in objects:
            DIRECT[object.label()] = dict()
            INVERSE[object.label()] = dict()
            DIRECT[object.label()]['data'] = dict()
            INVERSE[object.label()]['data'] = dict()
            INVERSE[object.label()]['percent'] = 77
        OBJECT_COUNT = 0
        PARAM_COUNT = 0

    CONTINUE = True

    if OBJECT_COUNT == N:
        CONTINUE = False
    if CONTINUE:
        object_count = 0
        for object in objects:
            label = object.label()
            param_count = 0
            for param in DISTANCE_FUNCTIONS:
                if object_count == OBJECT_COUNT and param_count == PARAM_COUNT :
                    DIRECT[label]['data'][param] = DISTANCE_FUNCTIONS[param](obj, object)
                    INVERSE[label]['data'][param] = 0
                    messages.info(request, f"{label}, {param} = {DIRECT[label]['data'][param]}", extra_tags='alert-secondary')
                    messages.info(request, f"[{OBJECT_COUNT}, {PARAM_COUNT}] из [{N}, {len(DISTANCE_FUNCTIONS)}]", extra_tags='alert-secondary')
                    PARAM_COUNT += 1
                    if PARAM_COUNT == len(DISTANCE_FUNCTIONS):
                        PARAM_COUNT = 0
                        OBJECT_COUNT += 1
                    context['continue'] = 'CONTINUE'
                    return context
                param_count += 1
            object_count += 1

    labels = []
    datasets = None
    legend = None
    sums = {}
    revs = {}
    nearest = None
    curr = 0

    for object in objects:
        label = object.label()
        labels.append(label)
        if not legend:
            legend = []
            datasets = []
            for param in DISTANCE_FUNCTIONS:
                legend.append(param)
                dataset = []
                datasets.append(dataset)

        for param in DISTANCE_FUNCTIONS:
            if not param in sums:
                sums[param] = DIRECT[label]['data'][param] # * (N+1)
            else:
                sums[param] += DIRECT[label]['data'][param] # * (N+1)
            if not param in revs:
                if DIRECT[label]['data'][param] != 0:
                    revs[param] = 1 / DIRECT[label]['data'][param]
                else:
                    revs[param] = 1000000000
            else:
                if DIRECT[label]['data'][param] != 0:
                    revs[param] += 1 / DIRECT[label]['data'][param]
                else:
                    revs[param] = 1000000000

    if not datasets:
        messages.info(request, f"Не с кем сравнивать {object}, нет выбранных авторов с хотя-бы одним выбранным произведением.)", extra_tags='alert-warning')
        return context

    for object in objects:
        label = object.label()
        percent = 0
        j = 0
        for param in DIRECT[label]['data']:
            if DIRECT[label]['data'][param] == 0:
                INVERSE[label]['data'][param] = 100
                percent += 100
            else:
                INVERSE[label]['data'][param] = 100 / (DIRECT[label]['data'][param] * revs[param])
                percent += 100 / (DIRECT[label]['data'][param] * revs[param])

        for param in DIRECT[label]['data']:
            if sums[param] == 0:
                DIRECT[label]['data'][param] = 0
            else:
                DIRECT[label]['data'][param] = 100 * DIRECT[label]['data'][param] / sums[param]

            datasets[j].append(INVERSE[label]['data'][param])
            j += 1

        INVERSE[label]['percent'] = percent / len(datasets) # common percent for all params

        if percent > curr:
            curr = percent
            nearest = object

    context['inverse'] = INVERSE
    context['vector'] = vector(DIRECT, INVERSE)
    context['nearest'] = nearest
    context['histograms'] = tools.plot_histogram(labels, datasets, legend)
    text = obj.text()
    context['plots'] = tools.plots(text).items()
    context['text'] = text
    tokens = tools.tokenize(text)
    context['tokens'] = tokens
    most_common = tools.most_common(tokens)
    context['most_common'] = most_common
    return context

class AuthorDetail(LoginRequiredMixin, DetailView) :
    model = Author

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)

        obj = super().get_object()
        count = Writing.objects.filter(author=obj, active=True).count()
        if count == 0:
            messages.info(self.request, f"У {obj} нет выбранных произведений, выберите хотя-бы одно.", extra_tags='alert-danger')
            return context

        authors = Author.objects.exclude(id=obj.id).annotate(Count('writings'))
        authors = authors.exclude(writings__count=0)
        authors = authors.filter(active=True)

        context = get_detail_context(self.request, context, obj, authors)
        return context

# _______________________________________________________ WRITING ______________

class WritingUpdate(LoginRequiredMixin, UpdateView) :
    model = Writing
    form_class = WritingForm
    success_url = reverse_lazy('pushkin:author_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

    def form_valid(self, form):
        form.instance.save()
        # form.instance.save_stats()
        # form.instance.author.save_stats()
        if form.instance.is_sample:
            self.success_url = reverse('pushkin:writing_detail', args=[form.instance.id])
            form.instance.save()
        return super().form_valid(form)

    def get_initial(self):
        # Get the initial dictionary from the superclass method
        initial = super(WritingUpdate, self).get_initial()
        # Copy the dictionary so we don't accidentally change a mutable dict
        initial = initial.copy()
        initial['hidden'] = self.object.hidden
        initial['is_sample'] = self.object.is_sample

        return initial

def writing_delete(request, id=None):
    if not id:
        messages.info(request, "Delete, No ID", extra_tags='alert-warning')
    writing = Writing.objects.filter(id=id).first()
    if not writing:
        messages.info(request, f"Delete, No {id} writing", extra_tags='alert-warning')
        return redirect(reverse('pushkin:author_list'))
    writing.delete()
    return redirect(reverse('pushkin:author_list'))


class WritingCreate(LoginRequiredMixin, CreateView) :
    model = Writing
    form_class = WritingForm
    success_url = reverse_lazy('pushkin:author_list')

    def form_valid(self, form):
        form.instance.save()
        # form.instance.save_stats()
        # form.instance.author.save_stats()
        return super().form_valid(form)


class WritingList(LoginRequiredMixin, ListView) :
    model = Writing

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        authors = Author.objects.all()
        context['authors'] = authors
        return context


class WritingDetail(LoginRequiredMixin, DetailView) :
    model = Writing

    def get_context_data(self, **kwargs):
        global CONTINUE

        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)

        obj = super().get_object()

        writings = Writing.objects.exclude(id=obj.id)
        writings = writings.filter(active=True, hidden=False, is_sample=False)

        context = get_detail_context(self.request, context, obj, writings)
        return context


class WritingInfo(LoginRequiredMixin, DetailView) :
    model = Writing
    template_name = 'pushkin/writing_info.html'

    def get_context_data(self, **kwargs):
        global CONTINUE

        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)

        obj = super().get_object()

        text = obj.text()
        context['plots'] = tools.plots(text).items()

        context['text'] = text
        context['char_count'] = len(text)

        cleaned_text = tools.cleaned(text)
        context['cleaned_text'] =cleaned_text
        context['cleaned_char_count'] = len(cleaned_text)

        tokens = tools.tokenize(text)
        context['tokens'] = tokens
        context['token_count'] = len(tokens)

        word_tokens = word_tokenize(text)
        context['word_tokens'] = word_tokens
        context['word_token_count'] = len(word_tokens)

        sent_tokens = sent_tokenize(text)
        context['sent_tokens'] = sent_tokens
        context['sent_token_count'] = len(sent_tokens)

        freq_dist = nltk.FreqDist(tokens)
        most_common = list(freq_dist.most_common(tools.TOP_MOST_COMMON_NUM))
        context['freq_dist'] = freq_dist
        context['most_common'] = most_common

        bigrams = list(nltk.bigrams(tokens))
        context['bigrams'] = bigrams
        context['bigrams_count'] = len(bigrams)

        bigrams_freq_dist = nltk.FreqDist(bigrams)
        bigrams_most_common = list(bigrams_freq_dist.most_common(tools.TOP_MOST_COMMON_NUM))
        context['bigrams_freq_dist'] = bigrams_freq_dist
        context['bigrams_most_common'] = bigrams_most_common

        return context

class WritingRead(LoginRequiredMixin, DetailView) :
    model = Writing
    template_name = 'pushkin/writing_read.html'

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        return context

# _________________________________________________________ SAMPLE _____________

def sample_update(request, id):
    if id:
        sample = Sample.objects.filter(id=id).first()
        if sample:
            writing = sample.writing
            if writing:
                return redirect(reverse('pushkin:writing_update', args=[sample.writing.id]))
            messages.info(request, f"Sample {id} has no writing", extra_tags='alert-warning')
            return redirect(reverse('pushkin:author_list'))
        messages.info(request, f"Sample {id} not found", extra_tags='alert-warning')
        return redirect(reverse('pushkin:author_list'))
    messages.info(request, f"Sample detail has no arg id", extra_tags='alert-warning')
    return redirect(reverse('pushkin:author_list'))

def sample_delete(request, id=None):
    if not id:
        messages.info(request, "Delete, No ID", extra_tags='alert-warning')
    sample = Sample.objects.filter(id=id).first()
    if not sample:
        messages.info(request, f"Delete, No {id} Sample", extra_tags='alert-warning')
        return redirect(reverse('pushkin:author_list'))
    sample.delete()
    return redirect(reverse('pushkin:author_list'))

def sample_create(request) :
    user = request.user
    sample = Sample.objects.filter(user=user).first()
    if sample:
        return redirect(reverse('pushkin:writing_update', args=[sample.writing.id]))
    sample = Sample.objects.create(user=user)
    author = Author.objects.create(name=user.username, hidden=True)
    writing = Writing.objects.create(author=author, hidden=True, is_sample=True)
    sample.writing = writing
    sample.save()
    return redirect(reverse('pushkin:writing_update', args=[sample.writing.id]))

class SampleList(LoginRequiredMixin, ListView) :
    model = Sample

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        return context

def sample_detail(request, id):
    if id:
        sample = Sample.objects.filter(id=id).first()
        if sample:
            writing = sample.writing
            if writing:
                return redirect(reverse('pushkin:writing_detail', args=[sample.writing.id]))
            messages.info(request, f"Sample {id} has no writing", extra_tags='alert-warning')
            return redirect(reverse('pushkin:author_list'))
        messages.info(request, f"Sample {id} not found", extra_tags='alert-warning')
        return redirect(reverse('pushkin:author_list'))
    messages.info(request, f"Sample detail has no arg id", extra_tags='alert-warning')
    return redirect(reverse('pushkin:author_list'))

# _________________________________________________________ UTILS ______________


def read_file_into_string(filename):
    strings = []
    with open(filename) as f:
        strings.append(f.read())
    return ' '.join(strings)


def writing_info(writing):
    info = {}
    f = str(writing.file)
    info['file'] = f
    info['size'] = round(os.stat(f).st_size / 1024)
    text = read_file_into_string(f)
    info['text'] = text
    info['tokens'] = tools.tokenize(text)
    return info

# ______________________________________________________ STATISTICS ____________

MATRIX = []
ROW_COUNT = 0
COLUMN_COUNT = 0
# CONTINUE = False

def distance_matrix(request, objects, param='mendenhall'):
    global MATRIX
    global ROW_COUNT
    global COLUMN_COUNT
    global CONTINUE

    func = DISTANCE_FUNCTIONS[param]

    CONTINUE = True
    row_count = 0
    n = objects.count()
    for object in objects:
        if (row_count == ROW_COUNT) :
            column_count = 0
            for other in objects:

                if column_count == COLUMN_COUNT :
                    if row_count > column_count :
                        MATRIX[row_count][column_count] = MATRIX[column_count][row_count]
                    else:
                        distance = func(object, other)
                        MATRIX[row_count][column_count] = distance

                    if not (row_count == n and column_count == n):
                        messages.info(request, f'{object} ({ROW_COUNT}) <=> {other} ({COLUMN_COUNT})', extra_tags='alert-success')
                    break
                column_count += 1
        row_count += 1

    COLUMN_COUNT += 1
    if COLUMN_COUNT == n:
        COLUMN_COUNT = 0
        ROW_COUNT += 1
    if ROW_COUNT == n:
        CONTINUE = False


def plot_stats(request, mode, param):
    global MATRIX
    global ROW_COUNT
    global COLUMN_COUNT
    global CONTINUE

    if mode == 'authors':
        authors = Author.objects.filter(active=True).annotate(Count('writings'))
        objects = authors.exclude(writings__count=0)
    else:
        objects = Writing.objects.filter(active=True)

    n = objects.count()

    if not CONTINUE: # First call
        MATRIX = []
        for i in range(n):
            row = []
            for j in range(n):
                row.append(0)
            MATRIX.append(row)

        ROW_COUNT = 0
        COLUMN_COUNT = 0
    distance_matrix(request, objects, param)

    data = dict()
    data['objects'] = objects
    data['matrix'] = MATRIX
    data['param'] = param
    data['mode'] = mode
    if CONTINUE:
       data['continue'] = 'CONTINUE'
    else:
        if PLOT_FUNCTIONS[param]:
            name = tools.plot_array(objects, param)
            fn = os.path.join(settings.MEDIA_URL, 'graphs', name)
            data['plot'] = fn

        sum = 0
        SUM = 0
        reverse_matrix = []
        direct_matrix = []
        for i in range(n):
            row = []
            ROW = []
            for j in range(n):
                if MATRIX[i][j] == 0:
                    row.append(1000000000)
                    ROW.append(0)
                else:
                    row.append(1 / MATRIX[i][j])
                    ROW.append(MATRIX[i][j])
                    sum += 1 / MATRIX[i][j]
                    SUM += MATRIX[i][j]
            reverse_matrix.append(row)
            direct_matrix.append(ROW)

        k = 100 / sum
        K = 100 / SUM

        for i in range(n):
            for j in range(n):
                if MATRIX[i][j] == 0:
                    reverse_matrix[i][j] = 100
                else:
                    reverse_matrix[i][j] = k * reverse_matrix[i][j]
                direct_matrix[i][j] = K * direct_matrix[i][j]

        data['direct'] = direct_matrix
        data['reverse'] = reverse_matrix

        workbook = Workbook()
        sheet = workbook.active

        c = 1
        r = 1
        s = 5
        for object in objects:
            sheet.cell(row=1, column=c+1).value = object.label()
            sheet.cell(row=1+n+s, column=c+1).value = object.label()
            sheet.cell(row=r+1, column=1).value = object.label()
            sheet.cell(row=r+1+n+s, column=1).value = object.label()
            c += 1
            r += 1
        for i in range(n):
            for j in range(n):
                sheet.cell(row=i+2, column=j+2).value = direct_matrix[i][j]
                sheet.cell(row=i+2+n+s, column=j+2).value = reverse_matrix[i][j]
                sheet.cell(row=i+2, column=j+2).number_format = '0.00'
                sheet.cell(row=i+2+n+s, column=j+2).number_format = '0.00'
        # for i in range(n):
        #     sheet.cell(row=3+n, column=j+2).value = '=SUM[B2:B10]' # f'=SUM(R[-{n+1}]C:R[-1]C)'
        #     sheet.cell(row=3+n+n+s, column=j+2).value = f'=SUM(R[-{7}]C:R[-1]C)'
        #     sheet.cell(row=i+2, column=n+3).value = f'=SUM((RC[-{n+1}]:RC[-1]))'
        #     sheet.cell(row=i+2+n+2, column=n+3).value = f'=SUM((RC[-{n+1}]:RC[-1]))'

        name = f'{param}_matrixes.xlsx'
        filename = os.path.join(settings.MEDIA_ROOT, 'excel', name)
        workbook.save(filename)

        files = []
        file = {}
        file['url'] = os.path.join(settings.MEDIA_URL, 'excel', name)
        file['name'] = name
        file['title'] = f'Скачать {name}'
        files.append(file)
        data['files'] = files

    return render(request, 'pushkin/plot_stats.html', data)


# _______________________________________________________________ MARKOV _______

def markovify_author(request, id):
    author = Author.objects.filter(id=id).first()
    text = author.text()
    count = Writing.objects.filter(author=author, active=True).count()
    if count == 0:
        messages.info(request, f"Нет текста, возможно у автора не выбраны произведения", extra_tags='alert-warning')
        return redirect(reverse('pushkin:author_list'))
    text_model = markovify.Text(text)
    masterpiece = ''
    for j in range(25):
        i = 0
        masterpiece += '<p>'
        while i < 3 :
            line = text_model.make_sentence()
            if line and line != '':
                line = re.sub(r"\{\d+\}", "", line)
                line = re.sub(r"\d\d\d\d", "", line)
                line = re.sub(r"x x x", "", line)
                line = re.sub(r"I|V|X|L|M", "", line)
                line = f'<span>{line}</span> '
                i += 1
                masterpiece += line
        masterpiece += '</p>'
    return render(request, 'pushkin/markov.html', {'title': author.name, 'masterpiece': masterpiece})

def markovify_writing(request, id):
    writing = Writing.objects.filter(id=id).first()
    text = writing.text()
    text_model = markovify.Text(text)
    masterpiece = ''
    for j in range(5):
        i = 0
        while i < 15 :
            line = text_model.make_sentence()
            if line and line != '':
                line = re.sub(r"\{\d+\}", "", line)
                line = re.sub(r"\d\d\d\d", "", line)
                line = re.sub(r"x x x", "", line)
                line = re.sub(r"I|V|X|L|M", "", line)
                i += 1
                masterpiece += line
        masterpiece += '\r\n'
    return render(request, 'pushkin/markov.html', {'title': f'{writing.author}, {writing}', 'masterpiece': masterpiece})


# ________________________________________________________________ API _________

@login_required
def author_info(request, id, mode='all', w_re=None):
    author = Author.objects.filter(id=id).first()
    if author :
        writing_list = []
        if w_re:
            writings = Writing.objects.filter(author=author, title__iregex=w_re)
        else:
            writings = Writing.objects.filter(author=author)
        if mode == 'selected_only':
            writings = writings.filter(active=True)
        for writing in writings:
            writing_list.append(writing.serialize())
        info = {
            'writings': writing_list
        }
        return JsonResponse(info, status=200)
    return JsonResponse({'error': f'user {id} not found'}, status=404)

@login_required
def active_count(request):
    authors_count = Author.objects.filter(active=True).count()
    writings_count = Writing.objects.filter(active=True).count()
    info = {
            'a_count': authors_count,
            'w_count': writings_count,
        }
    return JsonResponse(info, status=200)


@login_required
def all_active(request, mode=''):
    Author.objects.filter(active=False).update(active=True)
    if mode == 'with_writings':
        Writing.objects.filter(active=False).update(active=True)
    return redirect(reverse('pushkin:author_list'))

@login_required
def all_passive(request, mode=''):
    Author.objects.filter(active=True).update(active=False)
    if mode == 'with_writings':
        Writing.objects.filter(active=True).update(active=False)
    return redirect(reverse('pushkin:author_list'))

@csrf_exempt
@login_required
def author_active(request):
    if request.method == 'POST' :
        data = json.loads(request.body)
        if data.get("id") is not None:
            id = data["id"]
            if data.get("mode") is not None:
                mode = data["mode"]
                author = Author.objects.filter(id=id).first()
                if author :
                    author.active = True
                    author.save()
                    if mode == 'with_writings':
                        Writing.objects.filter(author=author).update(active=True)
                    return HttpResponse(status=200)
                return HttpResponse(status=404)
    return HttpResponse(status=400)

@csrf_exempt
@login_required
def author_passive(request):
    if request.method == 'POST' :
        data = json.loads(request.body)
        if data.get("id") is not None:
            id = data["id"]
            if data.get("mode") is not None:
                mode = data["mode"]
                author = Author.objects.filter(id=id).first()
                if author :
                    author.active = False
                    author.save()
                    if mode == 'with_writings':
                        Writing.objects.filter(author=author).update(active=False)
                    return HttpResponse(status=200)
                return HttpResponse(status=404)
    return HttpResponse(status=400)

@csrf_exempt
@login_required
def writing_check_active(request):
    if request.method == 'POST' :
        data = json.loads(request.body)
        if data.get("id") is not None:
            id = data["id"]
            writing = Writing.objects.filter(id=id).first()
            if writing :
                writing.active = not writing.active
                writing.save()
                return HttpResponse(status=200)
            return HttpResponse(status=404)
        return HttpResponse(status=400)
    return HttpResponse(status=401)

@csrf_exempt
@login_required
def writing_set_selected_genre(request):
    if request.method == 'POST' :
        data = json.loads(request.body)
        if data.get("genre") is not None:
            genre = data["genre"]
            Writing.objects.filter(active=True).update(genre=genre)
            return HttpResponse(status=200)
        return HttpResponse(status=400)
    return HttpResponse(status=401)

@login_required
def delete_selected(request):
    writings = Writing.objects.filter(active=True)
    authors = Author.objects.filter(active=True)
    if request.method == 'POST' :
        # messages.info(request, f'Delete {writings}, {authors}', extra_tags='alert-danger')
        writings.delete()
        authors.delete()
        return redirect(reverse('pushkin:author_list'))
    return render(request, 'pushkin/confirm_delete_selected.html', {'writings': writings, 'authors': authors})

@csrf_exempt
def search(request):
    if request.method == 'POST' :
        found = set()
        mode = None
        if 'mode' in request.POST :
            mode = request.POST['mode']
        a_re = ''
        if 'a_re' in request.POST :
            a_re = request.POST['a_re']
        w_re = ''
        if 'w_re' in request.POST :
            w_re = request.POST['w_re']
        if a_re == '' and w_re == '' and mode == 'all':
            messages.info(request, f'Значение для поиска не задано', extra_tags='alert-warning')
            return redirect(reverse('pushkin:author_list'))
        if a_re == '' :
            authors = Author.objects.all()
        else:
            authors = Author.objects.filter(name__iregex=a_re)
        if w_re == '':
            writings = Writing.objects.filter(author__in=authors)
        else:
            writings = Writing.objects.filter(title__iregex=w_re, author__in=authors)
            for writing in writings:
                found.add(writing.author.id)
            authors = authors.filter(id__in=found)

        active_authors = authors.filter(active=True)
        for author in active_authors:
            found.add(author.id)
        if mode == 'selected_only':
            writings = writings.filter(active=True)
            for writing in writings:
                found.add(writing.author.id)
            authors = Author.objects.filter(id__in=found)

        context = dict()
        context['a_re'] = a_re
        context['open'] = 'open'
        context['w_re'] = w_re
        context['mode'] = mode
        context['object_list'] = authors
        context['writings'] = writings
        context['active_a'] = active_authors.count()
        context['active_w'] = writings.filter(active=True).count()
        context['genres'] = GENRE_CHOICES
        messages.info(request, f'Поиск выполнен {found} {authors} {active_authors}', extra_tags='alert-success')
        return render(request, 'pushkin/author_list.html', context)

    messages.info(request, f'400', extra_tags='alert-warning')
    return redirect(reverse('pushkin:author_list'))


#__________________________________________________________ LIBRARY ____________

def library_view(request):
    filename = os.path.join(settings.STATIC_ROOT, 'pushkin', 'library.data')
    f = open(filename, 'rb')
    library = pickle.load(f)
    f.close()
    return render(request, 'pushkin/library.html', {'library': library})

def library_info(request):
    filename = os.path.join(settings.STATIC_ROOT, 'pushkin', 'library.data')
    f = open(filename, 'rb')
    library = pickle.load(f)
    f.close()
    for author in library:
        folder = author['folder']
        for writing in author['writings']:
            fn = os.path.join(settings.MEDIA_ROOT, 'library', folder, writing['fname'])
            writing['path'] = fn
            if os.path.exists(fn):
                size = os.stat(fn).st_size
                writing['size'] = size
                if size == 0:
                    writing['status'] = 'EMPTY'
                else:
                    writing['status'] = 'GOOD'
            else:
                writing['status'] = 'NOFILE'
    return render(request, 'pushkin/library.html', {'library': library})

def library_clean(request):
    filename = os.path.join(settings.STATIC_ROOT, 'pushkin', 'library.data')
    f = open(filename, 'rb')
    library = pickle.load(f)
    f.close()
    for author in library:
        folder = author['folder']
        for writing in author['writings']:
            fn = os.path.join(settings.MEDIA_ROOT, 'library', folder, writing['fname'])
            size = os.stat(fn).st_size
            writing['size'] = size
            if size == 0:
                if os.path.exists(fn):
                    os.remove(fn)
                author['writings'].remove(writing)
            else:
                writing['status'] = 'GOOD'
    f = open(filename, 'wb')
    pickle.dump(library, f)
    f.close()
    return render(request, 'pushkin/library.html', {'library': library})

def library_load(request):
    filename = os.path.join(settings.STATIC_ROOT, 'pushkin', 'library.data')
    f = open(filename, 'rb')
    library = pickle.load(f)
    f.close()

    for person in library:
        name = person['name']
        folder = person['folder']
        author = Author.objects.filter(name=name).first()
        if not author:
            author = Author.objects.create(name=name)
        for entry in person['writings']:
            title = entry['name']
            file = os.path.join(settings.MEDIA_ROOT, 'library', folder, entry['fname'])
            if os.path.exists(file):
                writing = Writing.objects.filter(title=title).first()
                if not writing:
                    writing = Writing.objects.create(author=author, title=title, file=file)

    return redirect(reverse('pushkin:author_list'))









